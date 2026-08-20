"""Panel de reportes mensuales para el departamento de contabilidad.

Consulta PostgreSQL, agrupa los marcajes del mes por empleado y exporta
el desglose de horas ordinarias, horas extra al 50% y horas extra al 100%
en Excel (``.xlsx``) o CSV, listo para la liquidación de haberes conforme
a la Ley N.º 213. Además emite el comprobante digital de cada marcación
con hash de seguridad y proyecta el Aguinaldo Proporcional (13.º salario,
Ley N.º 6380/2019).
"""

from __future__ import annotations

import csv
import hashlib
import hmac
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import auth
import reglamento
from database import Database
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

HORAS_BASE_MENSUAL: float = 160.0
RECARGO_EXTRA_50: float = 1.5
RECARGO_EXTRA_100: float = 2.0
CLAVE_COMPROBANTE: str = "sistema-marcacion-paraguay-2026"

ENCABEZADO_RESUMEN: List[str] = [
    "Usuario",
    "Nombre completo",
    "Días trabajados",
    "Tardanzas",
    "Horas ordinarias",
    "Horas extra 50%",
    "Horas extra 100%",
    "Total horas",
]

ENCABEZADO_DETALLE: List[str] = [
    "Usuario",
    "Nombre completo",
    "Fecha",
    "Entrada",
    "Salida",
    "Feriado",
    "Tardanza",
    "Horas ordinarias",
    "Horas extra 50%",
    "Horas extra 100%",
]


def _sumar(marcajes: List[Dict], campo: str) -> timedelta:
    """Suma una columna de tipo INTERVAL de una lista de marcajes."""
    return sum((m[campo] or timedelta(0) for m in marcajes), timedelta(0))


def _fmt(duracion: timedelta) -> str:
    """Formatea un ``timedelta`` como ``HH:MM`` para planillas contables."""
    total = int(duracion.total_seconds())
    signo = "-" if total < 0 else ""
    total = abs(total)
    horas, resto = divmod(total, 3600)
    minutos = resto // 60
    return f"{signo}{horas:02d}:{minutos:02d}"


def _agrupar(marcajes: List[Dict]) -> Dict[int, Dict[str, Any]]:
    """Agrupa marcajes por empleado conservando el orden de la consulta."""
    grupos: Dict[int, Dict[str, Any]] = {}
    for m in marcajes:
        grupo = grupos.setdefault(
            m["user_id"],
            {"usuario": m["username"], "nombre": m["full_name"], "marcajes": []},
        )
        grupo["marcajes"].append(m)
    return grupos


def _fila_resumen(grupo: Dict[str, Any]) -> List[Any]:
    """Construye la fila de resumen mensual de un empleado."""
    registros = grupo["marcajes"]
    ordinarias = _sumar(registros, "horas_ordinarias")
    extra_50 = _sumar(registros, "horas_extra_50")
    extra_100 = _sumar(registros, "horas_extra_100")
    tardanzas = sum(1 for m in registros if m["es_tardanza"])
    return [
        grupo["usuario"],
        grupo["nombre"],
        len(registros),
        tardanzas,
        _fmt(ordinarias),
        _fmt(extra_50),
        _fmt(extra_100),
        _fmt(ordinarias + extra_50 + extra_100),
    ]


def _exportar_xlsx(grupos: Dict[int, Dict[str, Any]], ruta: Path) -> None:
    """Genera el libro Excel con las hojas Resumen y Detalle."""
    from openpyxl import Workbook

    libro = Workbook()
    resumen = libro.active
    resumen.title = "Resumen"
    resumen.append(ENCABEZADO_RESUMEN)
    for grupo in grupos.values():
        resumen.append(_fila_resumen(grupo))
    detalle = libro.create_sheet("Detalle")
    detalle.append(ENCABEZADO_DETALLE)
    for grupo in grupos.values():
        for m in grupo["marcajes"]:
            detalle.append(
                [
                    grupo["usuario"],
                    grupo["nombre"],
                    m["hora_entrada"].date().isoformat(),
                    m["hora_entrada"].strftime("%H:%M"),
                    m["hora_salida"].strftime("%H:%M") if m["hora_salida"] else "en curso",
                    "Sí" if m["es_feriado"] else "No",
                    "Sí" if m["es_tardanza"] else "No",
                    _fmt(m["horas_ordinarias"]),
                    _fmt(m["horas_extra_50"]),
                    _fmt(m["horas_extra_100"]),
                ]
            )
    libro.save(ruta)


def _exportar_csv(grupos: Dict[int, Dict[str, Any]], ruta: Path) -> None:
    """Genera el CSV (UTF-8 BOM) con el resumen mensual por empleado."""
    with open(ruta, "w", newline="", encoding="utf-8-sig") as archivo:
        escritor = csv.writer(archivo)
        escritor.writerow(ENCABEZADO_RESUMEN)
        for grupo in grupos.values():
            escritor.writerow(_fila_resumen(grupo))


def _firma_comprobante(registro_id: int, tipo: str, momento: datetime) -> str:
    """Calcula la firma SHA-256 del registro para la fidelidad legal."""
    clave = os.getenv("COMPROBANTE_CLAVE", CLAVE_COMPROBANTE)
    origen = f"{registro_id}|{tipo}|{momento.isoformat()}|{clave}"
    return hashlib.sha256(origen.encode("utf-8")).hexdigest()[:16].upper()


def comprobante_marcacion(registro_id: int, momento: datetime, tipo: str) -> str:
    """Genera el comprobante digital tipo ticket de una marcación.

    Args:
        registro_id: Identificador del marcaje en ``marcajes``.
        momento: Instante exacto registrado por el motor.
        tipo: ``ENTRADA`` o ``SALIDA``.

    Returns:
        String con el ticket legible para imprimir o conservar como
        prueba de fidelidad del registro.
    """
    tipo = tipo.upper()
    if tipo not in ("ENTRADA", "SALIDA"):
        raise ValueError("El tipo debe ser 'ENTRADA' o 'SALIDA'.")
    firma = _firma_comprobante(registro_id, tipo, momento)
    return (
        "==========================================\n"
        "  COMPROBANTE DE MARCACIÓN\n"
        "  Sistema de Marcación - Ley 213/93\n"
        "==========================================\n"
        f"  Tipo: {tipo}\n"
        f"  ID del registro: {registro_id}\n"
        f"  Fecha: {momento.strftime('%d/%m/%Y')}\n"
        f"  Hora exacta: {momento.strftime('%H:%M:%S')}\n"
        f"  Hash de seguridad: {firma}\n"
        "==========================================\n"
        "  Conservar este comprobante como prueba\n"
        "  de fidelidad del registro.\n"
    )


def verificar_comprobante(
    registro_id: int, momento: datetime, tipo: str, firma: str
) -> bool:
    """Verifica la firma de un comprobante contra su recálculo."""
    esperada = _firma_comprobante(registro_id, tipo.upper(), momento)
    return hmac.compare_digest(esperada, firma.upper())


def calcular_aguinaldo(db: Database, anio: int) -> List[Dict[str, Any]]:
    """Proyecta el Aguinaldo Proporcional (13.º salario) por empleado.

    Fórmula (Ley N.º 6380/2019): la doceava parte de la remuneración del
    año, compuesta por el salario mensual multiplicado por los meses
    trabajados más el valor de las horas extra acumuladas (50% y 100%).
    El valor de la hora ordinaria se estima con el divisor de
    ``HORAS_BASE_MENSUAL`` (160 horas).
    """
    inicio_anio = date(anio, 1, 1)
    fin_anio = date(anio, 12, 31)
    extras = {r["user_id"]: r for r in db.get_horas_extra_year(anio)}
    resultados: List[Dict[str, Any]] = []
    for usuario in db.list_users():
        salario = float(usuario["salario_mensual"] or 0)
        creado = usuario["created_at"]
        base = max(creado.date(), inicio_anio)
        if base > fin_anio:
            meses = 0
        else:
            meses = (fin_anio.year - base.year) * 12 + (fin_anio.month - base.month) + 1
        extra_50 = extras.get(usuario["id"], {}).get("extra_50") or timedelta(0)
        extra_100 = extras.get(usuario["id"], {}).get("extra_100") or timedelta(0)
        valor_hora = salario / HORAS_BASE_MENSUAL if salario else 0.0
        horas_50 = extra_50.total_seconds() / 3600
        horas_100 = extra_100.total_seconds() / 3600
        valor_extras = valor_hora * (horas_50 * RECARGO_EXTRA_50 + horas_100 * RECARGO_EXTRA_100)
        aguinaldo = (salario * meses + valor_extras) / 12
        resultados.append(
            {
                "usuario": usuario["username"],
                "nombre": usuario["full_name"],
                "salario_mensual": salario,
                "meses_trabajados": meses,
                "extra_50": extra_50,
                "extra_100": extra_100,
                "valor_extras": valor_extras,
                "aguinaldo": aguinaldo,
            }
        )
    return resultados


def resumen_consulta(
    db: Database, user: Dict[str, Any], fecha: date
) -> Dict[str, Any]:
    """Compone el resumen transparente que un empleado ve en su consulta.

    Consulta de solo lectura: marcas de la fecha indicada, horas extra
    acumuladas del mes y aguinaldo proporcional del año (Ley 6380/2019).
    El resultado es JSON-serializable para alimentar la web y la GUI.

    Args:
        db: Capa de persistencia conectada.
        user: Empleado autenticado por cédula (dict de ``users``).
        fecha: Día del historial que se desea inspeccionar.

    Returns:
        Diccionario con ``usuario``, ``nombre``, ``fecha``, ``marcas_dia``,
        ``extras_mes`` y ``aguinaldo`` (o ``None`` si aún no proyecta).
    """
    marcas = db.get_entries_by_date(user["id"], fecha)
    marcas_dia = [
        {
            "id": m["id"],
            "entrada": m["hora_entrada"].strftime("%H:%M:%S"),
            "salida": m["hora_salida"].strftime("%H:%M:%S") if m["hora_salida"] else None,
            "tardanza": bool(m["es_tardanza"]),
            "feriado": bool(m["es_feriado"]),
            "tolerancia_aplicada": bool(m.get("tolerancia_aplicada")),
            "condicion_climatica": m.get("condicion_climatica") or "",
            "ordinarias": _fmt(m["horas_ordinarias"] or timedelta(0)),
            "extra_50": _fmt(m["horas_extra_50"] or timedelta(0)),
            "extra_100": _fmt(m["horas_extra_100"] or timedelta(0)),
        }
        for m in marcas
    ]
    del_mes = [m for m in db.get_marcajes_month(fecha.year, fecha.month) if m["user_id"] == user["id"]]
    extra_50 = sum(
        ((m["horas_extra_50"] or timedelta(0)) for m in del_mes), timedelta(0)
    )
    extra_100 = sum(
        ((m["horas_extra_100"] or timedelta(0)) for m in del_mes), timedelta(0)
    )
    horas_50 = extra_50.total_seconds() / 3600
    horas_100 = extra_100.total_seconds() / 3600
    proyeccion = next(
        (a for a in calcular_aguinaldo(db, fecha.year) if a["usuario"] == user["username"]),
        None,
    )
    return {
        "usuario": user["username"],
        "nombre": user["full_name"],
        "vinculo": user.get("tipo_vinculo") or "Funcionario",
        "fecha": fecha.isoformat(),
        "marcas_dia": marcas_dia,
        "extras_mes": {
            "horas_50": horas_50,
            "horas_100": horas_100,
            "texto_50": _fmt(extra_50),
            "texto_100": _fmt(extra_100),
        },
        "aguinaldo": proyeccion,
    }


def resumen_historico(
    db: Database, user: Dict[str, Any], desde: date, hasta: date
) -> Dict[str, Any]:
    """Compone el historial completo de un empleado dentro de un rango.

    Consulta indexada sobre ``(user_id, hora_entrada)`` que devuelve el
    detalle de cada marca, el acumulado de horas extra 50%/100% del período
    y el aguinaldo devengado en esos meses (Ley N.º 6380/2019).

    Args:
        db: Capa de persistencia conectada.
        user: Empleado identificado por su cédula (username).
        desde: Primer día del período a inspeccionar.
        hasta: Último día del período (no puede ser anterior a ``desde``).

    Returns:
        Diccionario JSON-serializable con ``marcas``, ``extras_periodo`` y
        ``aguinaldo_periodo``.
    """
    if hasta < desde:
        raise ValueError("La fecha 'hasta' no puede ser anterior a 'desde'.")
    hoy = datetime.now().date()
    if hasta > hoy:
        raise ValueError(
            f"La fecha 'hasta' no puede superar el día de hoy ({hoy.isoformat()})."
        )
    marcajes = db.get_marcajes_rango(user["id"], desde, hasta)
    marcas = [
        {
            "id": m["id"],
            "fecha": m["hora_entrada"].strftime("%Y-%m-%d"),
            "entrada": m["hora_entrada"].strftime("%H:%M:%S"),
            "salida": m["hora_salida"].strftime("%H:%M:%S") if m["hora_salida"] else None,
            "tardanza": bool(m["es_tardanza"]),
            "feriado": bool(m["es_feriado"]),
            "incidencia": m.get("tipo_incidencia") or "",
            "tolerancia_aplicada": bool(m.get("tolerancia_aplicada")),
            "condicion_climatica": m.get("condicion_climatica") or "",
            "ordinarias": _fmt(m["horas_ordinarias"] or timedelta(0)),
            "extra_50": _fmt(m["horas_extra_50"] or timedelta(0)),
            "extra_100": _fmt(m["horas_extra_100"] or timedelta(0)),
        }
        for m in marcajes
    ]
    extra_50 = sum(
        ((m["horas_extra_50"] or timedelta(0)) for m in marcajes), timedelta(0)
    )
    extra_100 = sum(
        ((m["horas_extra_100"] or timedelta(0)) for m in marcajes), timedelta(0)
    )
    return {
        "usuario": user["username"],
        "nombre": user["full_name"],
        "vinculo": user.get("tipo_vinculo") or "Funcionario",
        "desde": desde.isoformat(),
        "hasta": hasta.isoformat(),
        "marcas": marcas,
        "extras_periodo": {
            "horas_50": extra_50.total_seconds() / 3600,
            "horas_100": extra_100.total_seconds() / 3600,
            "texto_50": _fmt(extra_50),
            "texto_100": _fmt(extra_100),
        },
        "aguinaldo_periodo": aguinaldo_periodo(db, user, desde, hasta),
    }


def aguinaldo_periodo(
    db: Database, user: Dict[str, Any], desde: date, hasta: date
) -> Dict[str, Any]:
    """Proyecta el aguinaldo devengado dentro del período consultado.

    Cuenta los meses calendario completos entre ``desde`` y ``hasta``
    (limitados por la fecha de alta del empleado) y valora las horas extra
    del período al recargo legal para componer la doceava parte.

    Returns:
        Diccionario con ``meses_periodo``, ``valor_extras`` y ``aguinaldo``.
    """
    salario = float(user["salario_mensual"] or 0)
    alta = user["created_at"].date()
    base = max(desde.replace(day=1), alta.replace(day=1))
    meses = 0
    cursor = base
    while cursor <= hasta:
        meses += 1
        cursor = _sumar_mes(cursor)
    extras = db.get_marcajes_rango(user["id"], desde, hasta)
    extra_50 = sum(
        ((m["horas_extra_50"] or timedelta(0)) for m in extras), timedelta(0)
    )
    extra_100 = sum(
        ((m["horas_extra_100"] or timedelta(0)) for m in extras), timedelta(0)
    )
    valor_hora = salario / HORAS_BASE_MENSUAL if salario else 0.0
    valor_extras = valor_hora * (
        extra_50.total_seconds() / 3600 * RECARGO_EXTRA_50
        + extra_100.total_seconds() / 3600 * RECARGO_EXTRA_100
    )
    return {
        "meses_periodo": meses,
        "valor_extras": valor_extras,
        "aguinaldo": (salario * meses + valor_extras) / 12,
    }


def _sumar_mes(fecha: date) -> date:
    """Retorna el primer día del mes siguiente a la fecha dada."""
    if fecha.month == 12:
        return fecha.replace(year=fecha.year + 1, month=1)
    return fecha.replace(month=fecha.month + 1)


def exportar_aguinaldo(
    db: Database, actor: Dict, anio: int, ruta: Optional[str] = None
) -> str:
    """Exporta la proyección de aguinaldo a un Excel para RRHH.

    Solo Administrador y Recursos Humanos pueden ejecutarla (RBAC).
    """
    auth.require_role(db, actor, auth.ROLES_REPORTES)
    datos = calcular_aguinaldo(db, anio)
    if ruta is None:
        ruta = Path("reportes") / f"aguinaldo_{anio:04d}.xlsx"
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Aguinaldo"
    hoja.append(
        [
            "Usuario",
            "Nombre completo",
            "Salario mensual (Gs.)",
            "Meses trabajados",
            "Horas extra 50%",
            "Horas extra 100%",
            "Valor horas extra (Gs.)",
            "Aguinaldo proporcional (Gs.)",
        ]
    )
    for dato in datos:
        hoja.append(
            [
                dato["usuario"],
                dato["nombre"],
                f"{dato['salario_mensual']:,.0f}",
                dato["meses_trabajados"],
                _fmt(dato["extra_50"]),
                _fmt(dato["extra_100"]),
                f"{dato['valor_extras']:,.0f}",
                f"{dato['aguinaldo']:,.0f}",
            ]
        )
    libro.save(ruta)
    return str(ruta)


def exportar_asistencia_mensual(
    db: Database,
    actor: Dict,
    anio: int,
    mes: int,
    formato: str = "xlsx",
    ruta: Optional[str] = None,
) -> str:
    """Exporta la asistencia mensual de todos los empleados.

    Solo Administrador y Recursos Humanos pueden ejecutar la exportación
    (validado con RBAC dentro de este módulo).

    Args:
        db: Capa de persistencia conectada.
        actor: Usuario autenticado que solicita el reporte.
        anio: Año del periodo a exportar.
        mes: Mes (1-12) del periodo a exportar.
        formato: ``xlsx`` o ``csv``.
        ruta: Ruta de salida; por defecto ``reportes/asistencia_AAAA-MM.ext``.

    Returns:
        Ruta absoluta del archivo generado.
    """
    auth.require_role(db, actor, auth.ROLES_REPORTES)
    if formato not in ("xlsx", "csv"):
        raise ValueError("Formato no soportado. Use 'xlsx' o 'csv'.")
    marcajes = db.get_marcajes_month(anio, mes)
    grupos = _agrupar(marcajes)
    if ruta is None:
        ruta = Path("reportes") / f"asistencia_{anio:04d}-{mes:02d}.{formato}"
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    if formato == "xlsx":
        _exportar_xlsx(grupos, ruta)
    else:
        _exportar_csv(grupos, ruta)
    return str(ruta)


def obtener_metricas_tardanzas(
    db: Database, desde: Optional[date] = None, hasta: Optional[date] = None
) -> List[Dict[str, Any]]:
    """Distribución diaria de llegadas tardías para el tablero analítico.

    Completa con ceros los días sin incidencias para que el gráfico refleje
    el mes completo del 1 hasta hoy.
    """
    hoy = date.today()
    desde = desde or date(hoy.year, hoy.month, 1)
    hasta = hasta or hoy
    filas = db.get_metricas_tardanzas(desde, hasta)
    mapa = {fila["fecha"]: int(fila["cantidad"]) for fila in filas}
    dias: List[Dict[str, Any]] = []
    dia = desde
    while dia <= hasta:
        dias.append({"fecha": dia.isoformat(), "cantidad": mapa.get(dia, 0)})
        dia += timedelta(days=1)
    return dias


def obtener_horas_extra_por_departamento(db: Database) -> List[Dict[str, Any]]:
    """Horas extra acumuladas al 50% y 100% agrupadas por departamento."""
    filas = db.get_horas_extra_por_departamento()
    return [
        {
            "departamento": fila["departamento"],
            "horas_50": round(float(fila["horas_50"] or 0.0), 2),
            "horas_100": round(float(fila["horas_100"] or 0.0), 2),
        }
        for fila in filas
    ]


def obtener_proyeccion_aguinaldos_totales(
    db: Database, anio: Optional[int] = None
) -> Dict[str, Any]:
    """Proyección del aguinaldo proporcional de toda la empresa en Guaraníes.

    Acumula mes a mes desde enero (Ley N.º 6380/2019): por cada empleado
    activo se devenga ``salario_mensual / 12`` por mes transcurrido.
    """
    if anio is None:
        anio = date.today().year
    filas = db.get_proyeccion_aguinaldos()
    meses_transcurridos = date.today().month
    por_departamento: Dict[str, Dict[str, float]] = {}
    for fila in filas:
        salario = float(fila["salario_mensual"] or 0.0)
        acumulado = salario / 12.0 * meses_transcurridos
        resumen = por_departamento.setdefault(
            fila["departamento"], {"acumulado": 0.0, "anual": 0.0}
        )
        resumen["acumulado"] += acumulado
        resumen["anual"] += salario
    return {
        "anio": anio,
        "meses_transcurridos": meses_transcurridos,
        "empleados": len(filas),
        "total_acumulado_g": round(
            sum(v["acumulado"] for v in por_departamento.values())
        ),
        "total_anual_g": round(sum(v["anual"] for v in por_departamento.values())),
        "por_departamento": {
            departamento: {
                "acumulado_g": round(resumen["acumulado"]),
                "anual_g": round(resumen["anual"]),
            }
            for departamento, resumen in por_departamento.items()
        },
    }


def _vacaciones_devengadas(antiguedad_anios: float, vinculo: str = "Funcionario") -> float:
    """Días de vacaciones devengados según vínculo y antigüedad.

    Escala progresiva de la función pública paraguaya: 12 días hasta los
    5 años de servicio, 20 días entre 5 y 10 años y 30 días desde los
    10 años en adelante (Art. 29 Res. 1307/2010). Los pasantes usan la
    licencia de 10 días hábiles del Art. 23 de la Res. 3028/2024.
    """
    if vinculo == "Pasante":
        return 10.0
    return float(reglamento._vacaciones_funcionario(antiguedad_anios))


def resumen_empleado(
    db: Database, user: Dict[str, Any], fecha: Optional[date] = None
) -> Dict[str, Any]:
    """Compone el tablero personal del Portal del Empleado.

    Reúne en un solo diccionario JSON-serializable los tres bloques que
    alimentan las tarjetas informativas del empleado:

    - Vacaciones devengadas, usufructuadas y disponibles en el año en curso
      (Art. 23 de la Res. 3028/2024).
    - Contador de permisos del mes en curso, discriminado por tipo
      (salud, exámenes, lactancia, etc. — Art. 25).
    - Marcas mensuales con horas ordinarias por día y el acumulado de horas
      extra al 50% y 100% (Ley N.º 213).

    Args:
        db: Capa de persistencia conectada.
        user: Empleado autenticado (dict de ``users``).
        fecha: Día de referencia; por defecto la fecha actual del servidor.

    Returns:
        Diccionario con ``vacaciones``, ``permisos_mes``, ``marcas_mes``,
        ``extras_mes`` y la lista ``permisos`` para los botones de PDF.
    """
    hoy = fecha or date.today()
    vinculo = user.get("tipo_vinculo") or "Funcionario"
    antiguedad = (hoy - user["created_at"].date()).days / 365.25
    justificaciones = [
        j for j in db.list_justificaciones() if j["usuario_id"] == user["id"]
    ]
    tipo_vacaciones = "Vacaciones" if vinculo == "Funcionario" else "Licencia de Pasante"
    vacaciones_usadas = sum(
        (j["fecha_fin"] - j["fecha_inicio"]).days + 1
        for j in justificaciones
        if j["tipo_permiso"] == tipo_vacaciones
        and j["fecha_inicio"].year == hoy.year
    )
    permisos_mes = [
        j
        for j in justificaciones
        if j["tipo_permiso"] != tipo_vacaciones
        and j["fecha_inicio"].year == hoy.year
        and j["fecha_inicio"].month == hoy.month
    ]
    detalle: Dict[str, int] = {}
    for j in permisos_mes:
        detalle[j["tipo_permiso"]] = detalle.get(j["tipo_permiso"], 0) + 1
    marcajes = [
        m
        for m in db.get_marcajes_month(hoy.year, hoy.month)
        if m["user_id"] == user["id"]
    ]
    extra_50 = sum(
        ((m["horas_extra_50"] or timedelta(0)) for m in marcajes), timedelta(0)
    )
    extra_100 = sum(
        ((m["horas_extra_100"] or timedelta(0)) for m in marcajes), timedelta(0)
    )
    return {
        "usuario": user["username"],
        "nombre": user["full_name"],
        "vinculo": vinculo,
        "antiguedad_anios": round(antiguedad, 1),
        "reglamento": (
            reglamento.REGLAMENTO_PASANTIA
            if vinculo == "Pasante"
            else reglamento.REGLAMENTO_INTERNO
        ),
        "vacaciones": {
            "devengadas": _vacaciones_devengadas(antiguedad, vinculo),
            "usadas": vacaciones_usadas,
            "disponibles": max(
                0, _vacaciones_devengadas(antiguedad, vinculo) - vacaciones_usadas
            ),
        },
        "permisos_mes": {"total": len(permisos_mes), "detalle": detalle},
        "disponibilidad": reglamento.disponibilidad_permisos(db, user, hoy),
        "marcas_mes": {
            "dias": [m["hora_entrada"].strftime("%d") for m in marcajes],
            "ordinarias": [
                round((m["horas_ordinarias"] or timedelta(0)).total_seconds() / 3600, 2)
                for m in marcajes
            ],
        },
        "extras_mes": {
            "horas_50": round(extra_50.total_seconds() / 3600, 2),
            "horas_100": round(extra_100.total_seconds() / 3600, 2),
        },
        "permisos": [
            {
                "id": j["id"],
                "tipo": j["tipo_permiso"],
                "inicio": j["fecha_inicio"].isoformat(),
                "fin": j["fecha_fin"].isoformat(),
                "aprobador": j["aprobador"],
            }
            for j in justificaciones
        ],
    }


def _canonico_permiso(justificacion: Dict[str, Any]) -> str:
    """Serie canónica de bytes que el PDF autentica con SHA-256."""
    return (
        f"EMPRESA|3028/2024|{justificacion['id']}|{justificacion['username']}|"
        f"{justificacion['tipo_permiso']}|{justificacion['fecha_inicio'].isoformat()}|"
        f"{justificacion['fecha_fin'].isoformat()}|{justificacion['aprobador']}"
    )


def generar_pdf_permiso(solicitud_id: int) -> str:
    """Genera el PDF oficial de un permiso aprobado (Res. 3028/2024).

    Documento formal con membrete institucional simulado,
    los datos del pasante o funcionario, el período del permiso, las
    firmas electrónicas del tutor y el hash SHA-256 de validación legal,
    que queda persistido en ``justificaciones.hash_legal``.

    Args:
        solicitud_id: Identificador de la justificación aprobada.

    Returns:
        Ruta absoluta del PDF generado dentro de ``reportes/``.

    Raises:
        ValueError: Si el permiso no existe o no puede componerse.
    """
    db = Database()
    db.ensure_database()
    db.connect()
    try:
        justificacion = next(
            (j for j in db.list_justificaciones() if j["id"] == solicitud_id),
            None,
        )
        if not justificacion:
            raise ValueError(f"El permiso #{solicitud_id} no existe.")
        empleado = db.get_user_by_id(justificacion["usuario_id"])
        tutor = db.get_user_by_id(justificacion["aprobado_por"])
        if not empleado or not tutor:
            raise ValueError("Datos del empleado o tutor incompletos.")
        hash_legal = hashlib.sha256(
            _canonico_permiso(justificacion).encode("utf-8")
        ).hexdigest()
        db.actualizar_hash_justificacion(solicitud_id, hash_legal)
    finally:
        db.cerrar()

    carpeta = Path("reportes")
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta = carpeta / f"permiso_{solicitud_id:04d}.pdf"

    estilos = getSampleStyleSheet()
    normal = ParagraphStyle(
        "Normal", parent=estilos["Normal"], fontSize=9.5, leading=13
    )
    institucion = ParagraphStyle(
        "Institucion",
        parent=estilos["Title"],
        fontSize=17,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1A56DB"),
        spaceAfter=2,
    )
    subtitulo = ParagraphStyle(
        "Subtitulo",
        parent=normal,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#6B7280"),
        spaceAfter=10,
    )
    titulo_documento = ParagraphStyle(
        "TituloDocumento",
        parent=normal,
        fontSize=13,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        spaceBefore=8,
        spaceAfter=16,
    )
    pie = ParagraphStyle(
        "Pie", parent=normal, fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor("#6B7280")
    )

    dias = (justificacion["fecha_fin"] - justificacion["fecha_inicio"]).days + 1
    filas: List[List[Any]] = [
        ["Empleado", empleado["full_name"]],
        ["Cédula / Usuario", empleado["username"]],
        ["Vínculo laboral", empleado.get("tipo_vinculo") or "Funcionario"],
        ["Dependencia", empleado.get("departamento") or "General"],
        ["Tipo de permiso", justificacion["tipo_permiso"]],
        ["Período", f"{justificacion['fecha_inicio']} al {justificacion['fecha_fin']}"],
        ["Días hábiles", str(dias)],
        ["Aprobado por (tutor)", tutor["full_name"]],
        ["Fecha de emisión", date.today().isoformat()],
    ]
    tabla = Table(filas, colWidths=[45 * mm, 115 * mm])
    tabla.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6B7280")),
                ("TEXTCOLOR", (1, 0), (1, -1), colors.black),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E4E7EB")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    documento = SimpleDocTemplate(
        str(ruta), pagesize=A4, topMargin=20 * mm, bottomMargin=16 * mm
    )
    elemento = [
        Paragraph("Sistema de Marcación · Gestión de Asistencia", institucion),
        Paragraph("Dirección de Talento Humano", subtitulo),
        Paragraph(
            "Permiso aprobado en el marco de la Resolución de Directorio "
            "N.º 3028/2024 · Artículos 23 y 25",
            subtitulo,
        ),
        Paragraph(
            f"CERTIFICADO DE PERMISO OFICIAL N.º 3028-2024/{solicitud_id:04d}",
            titulo_documento,
        ),
        tabla,
        Spacer(1, 14 * mm),
        Paragraph(
            "____________________________________", ParagraphStyle("Firma", parent=normal, alignment=TA_CENTER)
        ),
        Paragraph(f"Firma electrónica del tutor · {tutor['full_name']}", normal),
        Spacer(1, 6 * mm),
        Paragraph("____________________________________", ParagraphStyle("Firma2", parent=normal, alignment=TA_CENTER)),
        Paragraph("Sello institucional · Dirección de Talento Humano", normal),
        Spacer(1, 12 * mm),
        Paragraph(
            f"Hash SHA-256 de validación legal: {hash_legal}", ParagraphStyle("Hash", parent=normal, fontSize=8)
        ),
        Paragraph(
            "Documento generado electrónicamente por el Sistema de Marcación. "
            "Cualquier alteración invalida el hash de autenticidad.",
            pie,
        ),
    ]
    documento.build(elemento)
    return str(ruta.absolute())